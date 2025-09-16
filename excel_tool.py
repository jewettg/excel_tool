#!/usr/bin/env python3
#
# Excel Spreadsheet Manipulation Script - Split Sheet
#
# Code written by:
#   Greg Jewett, jewettg@austin.utexas.edu, 512-471-9645
#
# Code maintained by:
#   Greg Jewett, jewettg@austin.utexas.edu, 512-471-9645
#
# Using pandas libraries, read in an Excel spreadsheet and perform related functions.
#
# ---------------------------------------------------------------------------------------
# CHANGE LOG
# 2025-09-10 (GSJ) Initial version.
#
# ---------------------------------------------------------------------------------------

# =======================================================================================
# BEGIN Import modules and dependencies
# =======================================================================================

# The system module for Python, specifically use to get command line arguments.
import sys

# =======================================================================================
# CHECK PYTHON VERSION
# Error out for any Python version earlier than minimum supported version.
# =======================================================================================
minVer = (3,11,0)
curVer = sys.version_info[0:]
if curVer < minVer:
    print("Current Python version: {}.{}.{}".format(*curVer+(0,0,)))
    print("ABORT: Expect Python version {}.{}.{}".format(*minVer+(0,0,))+" or better required!")
    sys.exit(1)

# Import Regular Expression method
import re

# The operating system module/library
import os

# Import the logging module, configuration in __main__
import logging

# Import the object-oriented filesystem paths "pathlib"
import pathlib

# Import time module to allow the script to "sleep"
import time

# YAML library for handling of YAML data types.
import yaml

# JSON library for handling of JSON data types.
import json

# Manipulation of date/time formats and data.
import datetime
from datetime import timezone
import pytz

# Pandas library for handling Excel spreadsheets.
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Add support for argument passing, used to collect information on what
# action to perform.
import argparse

# =======================================================================================
# END Import modules and dependencies
# =======================================================================================


# =======================================================================================
# BEGIN Functions used by global path variables.
# =======================================================================================
# Get Script Path
def scriptPath():
    return os.path.dirname(os.path.realpath(__file__))

# function to return a string (date/time) stamp, based on format needed.
def dt_stamp(format):
    stamp = datetime.datetime.now()
    if format == "d":
        # current date in ISO (YYYY-MM-DD) format
        return stamp.strftime("%Y-%m-%d")
    if format == "dt":
        # current date/time in ISO format: YYYY-MM-DDTHH:MM:SS.ddddd
        return stamp.isoformat()
    if format == "t":
        # current time in ISO format: HH:MM:SS.ddddd
        return stamp.strftime("%H:%M:%S.%f")
    if format == "fdt":
        # current date and time in format supported by OS for filenames.
        return stamp.strftime("%Y-%m-%d_%H%M%S")

# =======================================================================================
# END Functions used by global path variables.
# =======================================================================================

# =======================================================================================
# BEGIN Required variables and setup
# =======================================================================================

# Logging Metadata
scriptVer = "1.0"
scriptName = "Excel Manipulation Tools"
logTag = "EXCELTOOL"
logName = "excel_tools"
# logPath = "/opt/lb-bkups/script_logs/"+logName
logPath = scriptPath()+"/logs/"+logName
minLogLevel = logging.INFO

default_config_file = scriptPath()+"/excel_tool_config.yml"


# =======================================================================================
# END Required setup and global variables
# =======================================================================================




# =======================================================================================
# BEGIN Class Declarations
# =======================================================================================

# Define a class that will hold the request issued, along with status, etc..
# ------------------------------------------------------------------
class ExcelManipulationTool():

    # Class Methods
    # -------------------------------------------------------

    # Set the renewal status and if present, error message.
    def setStatus(self, status):
        if type(status) != bool:
            self.requestStatus = False
        else:
            self.requestStatus = status

    # Return the renewal status
    def getStatus(self):
        if hasattr(self, 'requestStatus'):
            if type(self.requestStatus) != bool:
                self.requestStatus = False
            else:
                return self.requestStatus
        else:
            return False



    def process_parameters(self):
        doLog.info("Processing command line parameters...")

        scriptDesc = "Excel Manipulation Tool - Perform various Excel spreadsheet manipulations."
        scriptContact = "Please contact Greg Jewett for support; greg@ejewett.com"
        aParser = argparse.ArgumentParser(  description = scriptDesc,
                                            epilog = scriptContact,
                                            add_help = True,
                                            allow_abbrev = False)


        subParsers = aParser.add_subparsers(help = 'sub-command help',
                                            required = True,
                                            dest = 'command')


        # ---------------------------------------------------------------------------------------
        # BEGIN SUB-COMMAND: SPLIT
        # ---------------------------------------------------------------------------------------
        theHelp = "Split an Excel file into multiple sheets.  Use 'split -h' to list of parameters"
        splitParser = subParsers.add_parser("split", help = theHelp)


        # PARAMETER:  Get the filename that contains PEM data to decode
        # ------------------------------------------------------------------------
        theHelp = "The path and filename of the excel file to split."
        splitParser.add_argument( "-f",
                                action = "store",
                                default = "",
                                type = str,
                                dest = "excelFile",
                                help = theHelp,
                                required = True)

        # PARAMETER:  Name of column to split on
        # ------------------------------------------------------------------------
        theHelp = "The name of the column to split the sheet on. "
        splitParser.add_argument( "-c",
                                action = "store",
                                type = str,
                                dest = "split_column",
                                help = theHelp,
                                required = True)
        # ---------------------------------------------------------------------------------------
        # END SUB-COMMAND: DECODE certificate PEM data
        # ---------------------------------------------------------------------------------------




        # Check to see if any parameters were provided, as there are some that are required.
        # If none provided, then output the help section.
        # ------------------------------------------------------------------------
        if len(sys.argv) < 2:
            aParser.print_help()
            sys.exit(1)


        # Setup class variables based on parameters
        # -------------------------------------------------------
        params = vars(aParser.parse_args())
        self.excel_file = params.get("excelFile")
        self.command = params.get("command")
        self.split_column = params.get("split_column")


    # SPLIT COMMAND
    # Split an Excel file into multiple sheets based on the agent name value.
    # ------------------------------------------------------------------------
    def split_sheet(self):
        doLog.info("Splitting Excel file: "+self.excel_file)

        if not os.path.isfile(self.excel_file):
            doLog.error("Excel file does not exist: "+self.excel_file)
            self.setStatus(False)
            return

        try:
            xls = pd.ExcelFile(self.excel_file)
        except Exception as e:
            doLog.error("Error reading Excel file: "+self.excel_file)
            doLog.error(e)
            self.setStatus(False)
            return

        # Create the "split" output directory if it does not already exist.
        output_dir = os.path.splitext(self.excel_file)[0]+"_split"
        pathlib.Path(output_dir).mkdir(exist_ok=True)

        sheet_names = xls.sheet_names
        doLog.info("Excel file contains the following sheets:")
        for sheet in sheet_names:
            doLog.info(" -> "+sheet)

        # Read in the first sheet only.
        try:
            df = pd.read_excel(self.excel_file, sheet_name=sheet_names[0])
        except Exception as e:
            doLog.error("Error reading first sheet of Excel file: "+self.excel_file)
            doLog.error(e)
            self.setStatus(False)
            return

        if self.split_column not in df.columns:
            doLog.error("Excel sheet does not contain '"+self.split_column+"' column, cannot split!")
            self.setStatus(False)
            return

        # Grab the first row (header row) to be used for all output files.
        header_row = pd.DataFrame([df.columns])  # DataFrame as single row with column names


        # Sort by specified column to ensure all like values are together.
        df = df.sort_values(by=[self.split_column], ascending=True)

        agent_names = df[self.split_column].unique()
        doLog.info("Found "+str(len(agent_names))+" unique "+self.split_column+" values to split on!")
        # for agent in agent_names:
        #     doLog.info("   - "+str(agent))

        # Split the dataframe based on unique values for column specified and write to separate files.
        for agent in agent_names:
            agent_df = df[df[self.split_column] == agent]
            clean_agent = re.sub(r'[^A-Za-z0-9_]', '_', str(agent)).replace('__', '_').strip('_')
            output_file = f"{os.path.splitext(self.excel_file)[0]}_{clean_agent}.xlsx"
            output_file = os.path.join(output_dir, os.path.basename(output_file))


            try:
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    header_row.to_excel(writer, index=False, header=False, startrow=0)
                    agent_df.to_excel(writer, index=False, header=False, startrow=1)
                doLog.info(f"Wrote {len(agent_df)} records to file: {output_file}")
            except Exception as e:
                doLog.error(f"Error writing Excel file for "+self.split_column+" '{agent}': {output_file}")
                doLog.error(e)
                self.setStatus(False)

            # Adjust column widths for all output files to fit content of all columns.
            # ------------------------------------------------------------------------
            # Load the workbook and select the active sheet
            workbook = load_workbook(output_file)
            sheet = workbook.active

            # Adjust column widths based on the header_row and agent_df
            for col in sheet.columns:
                max_length = 0
                column_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', etc.)
                for cell in col:
                    try:
                        if cell.value:  # Check if the cell has a value
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max_length + 2  # Add some padding
                sheet.column_dimensions[column_letter].width = adjusted_width  # Set the column width

            # Define the yellow fill style
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            # Apply the yellow fill to each cell in the header row
            for cell in sheet[1]:  # The header row is the first row (row 1)
                cell.fill = yellow_fill


            # Save the workbook after adjusting the widths
            workbook.save(output_file)



    # Initialization Method
    # -------------------------------------------------------
    def __init__(self, config_file):

        self.requestStatus = True
        self.config_file = config_file


        # Obtain default values from configuration (YML) file.
        # -------------------------------------------------------
        configData = dict()
        doLog.info("Reading configuration file: "+str(self.config_file))

        try:
            with open(self.config_file, "r") as configFile:
                configData = yaml.safe_load(configFile)
        except FileNotFoundError as error:
            doLog.error("File not found error reading configuration file: "+self.config_file)
            doLog.error(error)
            self.setStatus(False)
        except yaml.YAMLError as error:
            doLog.error("YAML error reading configuration file: "+self.config_file)
            doLog.error(error)
            self.setStatus(False)
        except IOError as error:
            doLog.error("I/O error reading configuration file: "+self.config_file)
            doLog.error(error)
            self.setStatus(False)
        except Exception as unknownError:
            doLog.error("Unknown error reading configuration file: "+self.config_file)
            doLog.error(unknownError)
            self.setStatus(False)


        self.process_parameters()


        # Setup class variables based on configuration file
        # -------------------------------------------------------
        # self.default_cert_org = configData.get("default_cert_org", "")
        # self.default_cert_state = configData.get("default_cert_state", "")



# =======================================================================================
# END Class Declarations
# =======================================================================================


# =======================================================================================
# BEGIN Functions Declarations
# =======================================================================================


# Setup and configure the logging system.
# -------------------------------------------------------
def setupLogging(logPath, logName, minLogLevel=logging.info, logTag="",
                 scriptName=logName, scriptVer="1.0", screenOut=False):

    logFile = logPath+"/"+logName+"_"+dt_stamp('d')+".log"
    levelOut = {50:"CRITICAL", 40:"ERROR", 30:"WARNING", 20:"INFO", 10:"DEBUG", 0:"NOTSET" }

    # Check if the logPath exists, if not create it.
    pathlib.Path(logPath).mkdir(parents=True, exist_ok=True)

    if screenOut:
        handlers = [logging.FileHandler(filename=logFile, mode='a'),
                    logging.StreamHandler(sys.stdout) ]
    else:
        handlers = [logging.FileHandler(filename=logFile, mode='a') ]

    logTag = (logTag if len(logTag) > 0 else re.sub("[^A-Z0-9]", "", re.sub("[a-z]", '', scriptName))[0:5])

    # Create the logging instance
    doLog = logging
    doLog.basicConfig(level=minLogLevel,
                      format="%(asctime)s %(levelname)-8s ("+logTag+") %(message)s",
                      handlers = handlers,
                      force=True)

    # Write out a log header
    doLog.info("----------------------------------------------------------------------")
    doLog.info(scriptName+"    v"+scriptVer)
    doLog.info("----------------------------------------------------------------------")
    doLog.info("Log Start: "+dt_stamp('d')+" "+dt_stamp('t'))
    doLog.info("Writing script log data to: "+logFile)

    doLog.info("Minimum logging level will be set to: "+levelOut.get(minLogLevel, "UNKNOWN"))
    doLog.disable(level=(minLogLevel-10))
    return doLog

# Display content in a list in columns
def display_columns(thelist, columns):
    rows = (len(thelist) + columns - 1) // columns
    formatted_rows = []
    for i in range(rows):
        row = ""
        for j in range(columns):
            index = i + j * rows
            if index < len(thelist):
                row += f"{thelist[index]:<30} "
        formatted_rows.append(row.strip())
    return formatted_rows



# =======================================================================================
# END Functions Declarations
# =======================================================================================




# =======================================================================================
# BEGIN Script
# =======================================================================================

if __name__ == '__main__':
    # Setup logging, instance is global, can be used by all functions.
    doLog = setupLogging(logPath, logName, minLogLevel, logTag, scriptName, scriptVer, screenOut=True)

    # Define "excelReq", an instance of ExcelManipulationTool
    excelReq = ExcelManipulationTool(default_config_file)

    if excelReq.getStatus():
        doLog.info("Starting Excel Manipulation Tool script...")
        doLog.info("Excel file to manipulate: "+str(excelReq.excel_file))

        match excelReq.command:
            case 'split':
                # Perform Excel manipulations
                excelReq.split_sheet()

    # ALL DONE, announce with status!
    if excelReq.getStatus():
        doLog.info("Script Finished!")
        doLog.info("All tasks completed successfully.")
    else:
        doLog.error("Script encountered errors in processing Excel manipulations!")
        doLog.error("One or more tasks failed, please check the log file for details.")

    # Shutdown the logger.
    doLog.shutdown()
# =======================================================================================
# END Script
# =======================================================================================
